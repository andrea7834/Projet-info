# -*- coding: utf-8 -*-

from PyQt5 import QtCore, QtGui, QtWidgets, Qt
import sys
from Saison import Saison
import pandas as pd
import numpy as np
from openpyxl import load_workbook


class Ui_La_Ligue_1_Uber_Eats(QtWidgets.QMainWindow):
    def setupUi(self, La_Ligue_1_Uber_Eats):
        self.name = La_Ligue_1_Uber_Eats
        saison = Saison()
        La_Ligue_1_Uber_Eats.setObjectName("La_Ligue_1_Uber_Eats")
        La_Ligue_1_Uber_Eats.resize(1183, 857)
        icon = QtGui.QIcon()
        icon.addPixmap(QtGui.QPixmap("C:\\Projet-info\\mini-soccer-petit-ballon-de-foot.jpg"), QtGui.QIcon.Normal,
                       QtGui.QIcon.Off)
        La_Ligue_1_Uber_Eats.setWindowIcon(icon)

        La_Ligue_1_Uber_Eats.setStyleSheet("background-color: rgb(134, 230, 96);\n" 
                                           "font: 75 12pt \"Times New Roman\";\n" "color: rgb(0, 125, 92);\n" 
                                           "border-color: rgb(0, 0, 0);")
        self.centralwidget = QtWidgets.QWidget(La_Ligue_1_Uber_Eats)
        self.centralwidget.setObjectName("centralwidget")
        self.main_layout = QtWidgets.QVBoxLayout(self.centralwidget)

        self.titre = QtWidgets.QLabel("La Ligue 1 UberEats", self)
        self.titre.move(450, 300)
        self.titre.resize(100, 50)
        self.titre.setAlignment(QtCore.Qt.AlignBottom)
        self.titre.setStyleSheet("border: 1px solid black;\n" "border-color: rgb(0, 125, 92);")

        self.top_layout = QtWidgets.QHBoxLayout()
        self.logo = QtGui.QPixmap("mini-soccer-petit-ballon-de-foot1.jpg")
        self.max_width = 200
        self.scaled_logo = self.logo.scaledToWidth(self.max_width)
        self.logo_label = QtWidgets.QLabel()
        self.logo_label.setPixmap(self.scaled_logo)
        self.logo_label.setAlignment(QtCore.Qt.AlignLeft)
        self.top_layout.addWidget(self.logo_label)
        self.main_layout.addLayout(self.top_layout)

        self.pixmap = QtGui.QPixmap("Logo_Ligue_1_Uber_Eats_2022.svg.png")
        self.max_width = 200
        self.scaled_pixmap = self.pixmap.scaledToWidth(self.max_width)
        self.image_label = QtWidgets.QLabel()
        self.image_label.setPixmap(self.scaled_pixmap)
        self.image_label.setAlignment(QtCore.Qt.AlignRight)
        # self.image_label.setGeometry(1200, 10, 300, 200)
        self.top_layout.addWidget(self.image_label)
        self.main_layout.addLayout(self.top_layout)

        self.saison = Saison()
        self.classement_final = saison.fin
        # self.classement_final = pd.DataFrame(data=saison.classement_final)
        self.resultats = QtWidgets.QTableWidget(self.centralwidget)
        self.resultats.setGeometry(QtCore.QRect(25, 400, 1133, 337))
        self.resultats.setStyleSheet("background-color: rgb(255, 255, 255);")
        self.resultats.setObjectName("resultats")
        self.resultats.setHorizontalHeaderLabels(("Clubs", 'Points'))
        self.resultats.setColumnCount(2)
        self.resultats.setRowCount(20)
        self.resultats.setColumnWidth(0, 540)
        self.resultats.setColumnWidth(1, 540)
        for ligne in range(20):
            self.resultats.setItem(ligne, 0, Qt.QTableWidgetItem(str(self.classement_final['Clubs'][ligne])))
            self.resultats.setItem(ligne, 1, Qt.QTableWidgetItem(str(self.classement_final['Points'][ligne])))

        self.instructions = QtWidgets.QTextEdit(self.centralwidget)
        self.instructions.setGeometry(QtCore.QRect(25, 225, 200, 90))
        self.instructions.setObjectName("instructions")

        self.titre1 = QtWidgets.QTextEdit(self.centralwidget)
        self.titre1.setGeometry(QtCore.QRect(365, 320, 450, 70))
        self.titre1.setStyleSheet("border: 1px solid black;\n" "border-color: rgb(134, 230, 96);")
        self.titre1.setObjectName("titre1")

        self.quitter = QtWidgets.QPushButton(self.centralwidget)
        self.quitter.setGeometry(QtCore.QRect(530, 775, 93, 28))
        self.quitter.setObjectName("quitter")
        self.menu_file_quit = QtWidgets.QAction('&Quit')
        self.menu_file_quit.setToolTip('Exit the application.')
        self.menu_file_quit.triggered.connect(self.quit)
        self.menu_file_quit.setShortcut(QtGui.QKeySequence("Ctrl+Q"))

        self.widget = QtWidgets.QWidget(self.centralwidget)
        self.widget.setGeometry(QtCore.QRect(300, 40, 250, 250))
        self.widget.setObjectName("widget")
        self.calendrier = QtWidgets.QGridLayout(self.widget)
        self.calendrier.setSizeConstraint(QtWidgets.QLayout.SetMinimumSize)
        self.calendrier.setContentsMargins(0, 0, 0, 0)
        self.calendrier.setObjectName("calendrier")

        self.jour1 = QtWidgets.QPushButton(self.widget)
        self.jour1.setStyleSheet("background-color: rgb(216, 216, 216);\n"
                                "font: 75 12pt \"Times New Roman\";\n"
                                "color: rgb(0, 0, 0);\n"
                                "background-color: rgb(255, 255, 255);")
        self.jour1.setObjectName("jour1")

        self.jour2 = QtWidgets.QPushButton(self.widget)
        self.jour2.setStyleSheet("font: 75 12pt \"Times New Roman\";\n"
"background-color: rgb(4, 4, 4);\n"
"color: rgb(255, 255, 255);")
        self.jour2.setObjectName("jour2")
        self.calendrier.addWidget(self.jour2, 0, 1, 1, 1)

        self.jour3 = QtWidgets.QPushButton(self.widget)
        self.jour3.setStyleSheet("background-color: rgb(216, 216, 216);\n"
"font: 75 12pt \"Times New Roman\";\n"
"color: rgb(0, 0, 0);\n"
"background-color: rgb(255, 255, 255);")
        self.jour3.setObjectName("jour3")
        self.calendrier.addWidget(self.jour3, 0, 2, 1, 1)

        self.jour4 = QtWidgets.QPushButton(self.widget)
        self.jour4.setStyleSheet("font: 75 12pt \"Times New Roman\";\n"
"background-color: rgb(4, 4, 4);\n"
"color: rgb(255, 255, 255);")
        self.jour4.setObjectName("jour4")
        self.calendrier.addWidget(self.jour4, 0, 3, 1, 1)

        self.jour5 = QtWidgets.QPushButton(self.widget)
        self.jour5.setStyleSheet("background-color: rgb(216, 216, 216);\n"
"font: 75 12pt \"Times New Roman\";\n"
"color: rgb(0, 0, 0);\n"
"background-color: rgb(255, 255, 255);")
        self.jour5.setObjectName("jour5")
        self.calendrier.addWidget(self.jour5, 0, 4, 1, 1)

        self.jour6 = QtWidgets.QPushButton(self.widget)
        self.jour6.setStyleSheet("font: 75 12pt \"Times New Roman\";\n"
"background-color: rgb(4, 4, 4);\n"
"color: rgb(255, 255, 255);")
        self.jour6.setObjectName("jour6")
        self.calendrier.addWidget(self.jour6, 0, 5, 1, 1)

        self.jour7 = QtWidgets.QPushButton(self.widget)
        self.jour7.setStyleSheet("font: 75 12pt \"Times New Roman\";\n"
"background-color: rgb(4, 4, 4);\n"
"color: rgb(255, 255, 255);")
        self.jour7.setObjectName("jour7")
        self.calendrier.addWidget(self.jour7, 1, 0, 1, 1)

        self.jour8 = QtWidgets.QPushButton(self.widget)
        self.jour8.setStyleSheet("background-color: rgb(216, 216, 216);\n"
"font: 75 12pt \"Times New Roman\";\n"
"color: rgb(0, 0, 0);\n"
"background-color: rgb(255, 255, 255);")
        self.jour8.setObjectName("jour8")
        self.calendrier.addWidget(self.jour8, 1, 1, 1, 1)

        self.jour9 = QtWidgets.QPushButton(self.widget)
        self.jour9.setStyleSheet("font: 75 12pt \"Times New Roman\";\n"
"background-color: rgb(4, 4, 4);\n"
"color: rgb(255, 255, 255);")
        self.jour9.setObjectName("jour9")
        self.calendrier.addWidget(self.jour9, 1, 2, 1, 1)

        self.jour10 = QtWidgets.QPushButton(self.widget)
        self.jour10.setStyleSheet("background-color: rgb(216, 216, 216);\n"
"font: 75 12pt \"Times New Roman\";\n"
"color: rgb(0, 0, 0);\n"
"background-color: rgb(255, 255, 255);")
        self.jour10.setObjectName("jour10")
        self.calendrier.addWidget(self.jour10, 1, 3, 1, 1)

        self.jour11 = QtWidgets.QPushButton(self.widget)
        self.jour11.setStyleSheet("font: 75 12pt \"Times New Roman\";\n"
"background-color: rgb(4, 4, 4);\n"
"color: rgb(255, 255, 255);")
        self.jour11.setObjectName("jour11")
        self.calendrier.addWidget(self.jour11, 1, 4, 1, 1)

        self.jour12 = QtWidgets.QPushButton(self.widget)
        self.jour12.setStyleSheet("background-color: rgb(216, 216, 216);\n"
"font: 75 12pt \"Times New Roman\";\n"
"color: rgb(0, 0, 0);\n"
"background-color: rgb(255, 255, 255);")
        self.jour12.setObjectName("jour12")
        self.calendrier.addWidget(self.jour12, 1, 5, 1, 1)

        self.jour13 = QtWidgets.QPushButton(self.widget)
        self.jour13.setStyleSheet("background-color: rgb(216, 216, 216);\n"
"font: 75 12pt \"Times New Roman\";\n"
"color: rgb(0, 0, 0);\n"
"background-color: rgb(255, 255, 255);")
        self.jour13.setObjectName("jour13")
        self.calendrier.addWidget(self.jour13, 2, 0, 1, 1)

        self.jour14 = QtWidgets.QPushButton(self.widget)
        self.jour14.setStyleSheet("font: 75 12pt \"Times New Roman\";\n"
"background-color: rgb(4, 4, 4);\n"
"color: rgb(255, 255, 255);")
        self.jour14.setObjectName("jour14")
        self.calendrier.addWidget(self.jour14, 2, 1, 1, 1)

        self.jour15 = QtWidgets.QPushButton(self.widget)
        self.jour15.setStyleSheet("background-color: rgb(216, 216, 216);\n"
"font: 75 12pt \"Times New Roman\";\n"
"color: rgb(0, 0, 0);\n"
"background-color: rgb(255, 255, 255);")
        self.jour15.setObjectName("jour15")
        self.calendrier.addWidget(self.jour15, 2, 2, 1, 1)

        self.jour16 = QtWidgets.QPushButton(self.widget)
        self.jour16.setStyleSheet("font: 75 12pt \"Times New Roman\";\n"
"background-color: rgb(4, 4, 4);\n"
"color: rgb(255, 255, 255);")
        self.jour16.setObjectName("jour16")
        self.calendrier.addWidget(self.jour16, 2, 3, 1, 1)

        self.jour17 = QtWidgets.QPushButton(self.widget)
        self.jour17.setStyleSheet("background-color: rgb(216, 216, 216);\n"
"font: 75 12pt \"Times New Roman\";\n"
"color: rgb(0, 0, 0);\n"
"background-color: rgb(255, 255, 255);")
        self.jour17.setObjectName("jour17")
        self.calendrier.addWidget(self.jour17, 2, 4, 1, 1)

        self.jour18 = QtWidgets.QPushButton(self.widget)
        self.jour18.setStyleSheet("font: 75 12pt \"Times New Roman\";\n"
"background-color: rgb(4, 4, 4);\n"
"color: rgb(255, 255, 255);")
        self.jour18.setObjectName("jour18")
        self.calendrier.addWidget(self.jour18, 2, 5, 1, 1)

        self.jour19 = QtWidgets.QPushButton(self.widget)
        self.jour19.setStyleSheet("font: 75 12pt \"Times New Roman\";\n"
"background-color: rgb(4, 4, 4);\n"
"color: rgb(255, 255, 255);")
        self.jour19.setObjectName("jour19")
        self.calendrier.addWidget(self.jour19, 3, 0, 1, 1)

        self.jour20 = QtWidgets.QPushButton(self.widget)
        self.jour20.setStyleSheet("background-color: rgb(216, 216, 216);\n"
"font: 75 12pt \"Times New Roman\";\n"
"color: rgb(0, 0, 0);\n"
"background-color: rgb(255, 255, 255);")
        self.jour20.setObjectName("jour20")
        self.calendrier.addWidget(self.jour20, 3, 1, 1, 1)

        self.jour21 = QtWidgets.QPushButton(self.widget)
        self.jour21.setStyleSheet("font: 75 12pt \"Times New Roman\";\n"
"background-color: rgb(4, 4, 4);\n"
"color: rgb(255, 255, 255);")
        self.jour21.setObjectName("jour21")
        self.calendrier.addWidget(self.jour21, 3, 2, 1, 1)

        self.jour22 = QtWidgets.QPushButton(self.widget)
        self.jour22.setStyleSheet("background-color: rgb(216, 216, 216);\n"
"font: 75 12pt \"Times New Roman\";\n"
"color: rgb(0, 0, 0);\n"
"background-color: rgb(255, 255, 255);")
        self.jour22.setObjectName("jour22")
        self.calendrier.addWidget(self.jour22, 3, 3, 1, 1)
        self.jour23 = QtWidgets.QPushButton(self.widget)
        self.jour23.setStyleSheet("font: 75 12pt \"Times New Roman\";\n"
"background-color: rgb(4, 4, 4);\n"
"color: rgb(255, 255, 255);")
        self.jour23.setObjectName("jour23")
        self.calendrier.addWidget(self.jour23, 3, 4, 1, 1)
        self.jour24 = QtWidgets.QPushButton(self.widget)
        self.jour24.setStyleSheet("background-color: rgb(216, 216, 216);\n"
"font: 75 12pt \"Times New Roman\";\n"
"color: rgb(0, 0, 0);\n"
"background-color: rgb(255, 255, 255);")
        self.jour24.setObjectName("jour24")
        self.calendrier.addWidget(self.jour24, 3, 5, 1, 1)
        self.jour25 = QtWidgets.QPushButton(self.widget)
        self.jour25.setStyleSheet("background-color: rgb(216, 216, 216);\n"
"font: 75 12pt \"Times New Roman\";\n"
"color: rgb(0, 0, 0);\n"
"background-color: rgb(255, 255, 255);")
        self.jour25.setObjectName("jour25")
        self.calendrier.addWidget(self.jour25, 4, 0, 1, 1)
        self.jour26 = QtWidgets.QPushButton(self.widget)
        self.jour26.setStyleSheet("font: 75 12pt \"Times New Roman\";\n"
"background-color: rgb(4, 4, 4);\n"
"color: rgb(255, 255, 255);")
        self.jour26.setObjectName("jour26")
        self.calendrier.addWidget(self.jour26, 4, 1, 1, 1)
        self.jour27 = QtWidgets.QPushButton(self.widget)
        self.jour27.setStyleSheet("background-color: rgb(216, 216, 216);\n"
"font: 75 12pt \"Times New Roman\";\n"
"color: rgb(0, 0, 0);\n"
"background-color: rgb(255, 255, 255);")
        self.jour27.setObjectName("jour27")
        self.calendrier.addWidget(self.jour27, 4, 2, 1, 1)
        self.jour28 = QtWidgets.QPushButton(self.widget)
        self.jour28.setStyleSheet("font: 75 12pt \"Times New Roman\";\n"
"background-color: rgb(4, 4, 4);\n"
"color: rgb(255, 255, 255);")
        self.jour28.setObjectName("jour28")
        self.calendrier.addWidget(self.jour28, 4, 3, 1, 1)
        self.jour29 = QtWidgets.QPushButton(self.widget)
        self.jour29.setStyleSheet("background-color: rgb(216, 216, 216);\n"
"font: 75 12pt \"Times New Roman\";\n"
"color: rgb(0, 0, 0);\n"
"background-color: rgb(255, 255, 255);")
        self.jour29.setObjectName("jour29")
        self.calendrier.addWidget(self.jour29, 4, 4, 1, 1)
        self.jour30 = QtWidgets.QPushButton(self.widget)
        self.jour30.setStyleSheet("font: 75 12pt \"Times New Roman\";\n"
"background-color: rgb(4, 4, 4);\n"
"color: rgb(255, 255, 255);")
        self.jour30.setObjectName("jour30")
        self.calendrier.addWidget(self.jour30, 4, 5, 1, 1)
        self.jour31 = QtWidgets.QPushButton(self.widget)
        self.jour31.setStyleSheet("font: 75 12pt \"Times New Roman\";\n"
"background-color: rgb(4, 4, 4);\n"
"color: rgb(255, 255, 255);")
        self.jour31.setObjectName("jour31")
        self.calendrier.addWidget(self.jour31, 5, 0, 1, 1)
        self.jour32 = QtWidgets.QPushButton(self.widget)
        self.jour32.setStyleSheet("background-color: rgb(216, 216, 216);\n"
"font: 75 12pt \"Times New Roman\";\n"
"color: rgb(0, 0, 0);\n"
"background-color: rgb(255, 255, 255);")
        self.jour32.setObjectName("jour32")
        self.calendrier.addWidget(self.jour32, 5, 1, 1, 1)
        self.jour33 = QtWidgets.QPushButton(self.widget)
        self.jour33.setStyleSheet("font: 75 12pt \"Times New Roman\";\n"
"background-color: rgb(4, 4, 4);\n"
"color: rgb(255, 255, 255);")
        self.jour33.setObjectName("jour33")
        self.calendrier.addWidget(self.jour33, 5, 2, 1, 1)
        self.jour34 = QtWidgets.QPushButton(self.widget)
        self.jour34.setStyleSheet("background-color: rgb(216, 216, 216);\n"
"font: 75 12pt \"Times New Roman\";\n"
"color: rgb(0, 0, 0);\n"
"background-color: rgb(255, 255, 255);")
        self.jour34.setObjectName("jour34")
        self.calendrier.addWidget(self.jour34, 5, 3, 1, 1)
        self.jour35 = QtWidgets.QPushButton(self.widget)
        self.jour35.setStyleSheet("font: 75 12pt \"Times New Roman\";\n"
"background-color: rgb(4, 4, 4);\n"
"color: rgb(255, 255, 255);")
        self.jour35.setObjectName("jour35")
        self.calendrier.addWidget(self.jour35, 5, 4, 1, 1)
        self.jour36 = QtWidgets.QPushButton(self.widget)
        self.jour36.setStyleSheet("background-color: rgb(216, 216, 216);\n"
"font: 75 12pt \"Times New Roman\";\n"
"color: rgb(0, 0, 0);\n"
"background-color: rgb(255, 255, 255);")
        self.jour36.setObjectName("jour36")
        self.calendrier.addWidget(self.jour36, 5, 5, 1, 1)
        self.jour37 = QtWidgets.QPushButton(self.widget)
        self.jour37.setStyleSheet("background-color: rgb(216, 216, 216);\n"
"font: 75 12pt \"Times New Roman\";\n"
"color: rgb(0, 0, 0);\n"
"background-color: rgb(255, 255, 255);")
        self.jour37.setObjectName("jour37")
        self.calendrier.addWidget(self.jour37, 6, 0, 1, 1)
        self.jour38 = QtWidgets.QPushButton(self.widget)
        self.jour38.setStyleSheet("font: 75 12pt \"Times New Roman\";\n"
"background-color: rgb(4, 4, 4);\n"
"color: rgb(255, 255, 255);")
        self.jour38.setObjectName("jour38")
        self.calendrier.addWidget(self.jour38, 6, 1, 1, 1)
        La_Ligue_1_Uber_Eats.setCentralWidget(self.centralwidget)

        self.retranslateUi(La_Ligue_1_Uber_Eats)
        self.quitter.clicked.connect(La_Ligue_1_Uber_Eats.close)
        QtCore.QMetaObject.connectSlotsByName(La_Ligue_1_Uber_Eats)

        boutons = [self.jour1, self.jour2, self.jour3, self.jour4, self.jour5, self.jour6, self.jour7, self.jour8,
                   self.jour9, self.jour10, self.jour11, self.jour12, self.jour13, self.jour14, self.jour15, self.jour16,
                   self.jour17, self.jour18, self.jour19, self.jour20, self.jour21, self.jour22, self.jour23, self.jour24,
                   self.jour25, self.jour26, self.jour27, self.jour28, self.jour29, self.jour30, self.jour31, self.jour32,
                   self.jour33, self.jour34, self.jour35, self.jour36, self.jour37, self.jour38]
        for i in range(len(boutons)):
            boutons[i].clicked.connect(self.handle_button_clicked)
            boutons[i].setProperty('index', i + 1)
        # if boutons.any().clicked:
        #     indice = list(boutons).index(boutons.any().clicked)
        #     self.load_data(indice+1)

    def handle_button_clicked(self):
        sender = self.sender()  # Obtenir le bouton qui a émis le signal
        index = sender.property('index')  # Obtenir la propriété 'index' du bouton

        # Utilisation de l'indice pour appeler la fonction load_data
        self.load_data(index)


    def load_data(self, no_jour):
        # self.main_layout = QtWidgets.QVBoxLayout(self.centralwidget)

        self.workbook = load_workbook("jour{}.xlsx".format(no_jour))
        self.feuille = self.workbook.active

        self.resultats = QtWidgets.QTableWidget(self.centralwidget)
        self.resultats.setRowCount(10)
        self.resultats.setColumnCount(8)
        self.resultats.setGeometry(QtCore.QRect(25, 400, 1133, 337))
        self.resultats.setStyleSheet("background-color: rgb(255, 255, 255);")
        self.resultats.setObjectName("resultats")
        self.resultats.setHorizontalHeaderLabels(
                    ("Points dom", '"Buteurs dom"', "Clubs à domicile", "Buts dom", "Buts exté",
                     "Clubs à l'extérieur", 'Buteurs exté', 'Points exté'))

        self.resultats.setColumnWidth(0, 100)
        self.resultats.setColumnWidth(1, 175)
        self.resultats.setColumnWidth(2, 175)
        self.resultats.setColumnWidth(3, 100)
        self.resultats.setColumnWidth(4, 100)
        self.resultats.setColumnWidth(5, 175)
        self.resultats.setColumnWidth(6, 175)
        self.resultats.setColumnWidth(7, 100)

        for i in range(2, self.feuille.max_row+1):
            for j in range(2, self.feuille.max_column+1):
                cell_value = self.feuille.cell(row=i, column=j).value
                item = QtWidgets.QTableWidgetItem(str(cell_value))
                self.resultats.setItem(i-2 , j-2, item)
        # self.resultats.removeColumn(0)
        self.resultats.show()


    def quit(self):
        sys.exit()

    def retranslateUi(self, La_Ligue_1_Uber_Eats):
        _translate = QtCore.QCoreApplication.translate
        La_Ligue_1_Uber_Eats.setWindowTitle(_translate("La_Ligue_1_Uber_Eats", "La_Ligue_1_Uber_Eats"))
        self.instructions.setHtml(_translate("La_Ligue_1_Uber_Eats", "<!DOCTYPE HTML PUBLIC \"-//W3C//DTD HTML 4.0//EN\" \"http://www.w3.org/TR/REC-html40/strict.dtd\">\n"
"<html><head><meta name=\"qrichtext\" content=\"1\" /><style type=\"text/css\">\n"
"p, li { white-space: pre-wrap; }\n"
"</style></head><body style=\" font-family:\'Times New Roman\'; font-size:12pt; font-weight:72; font-style:normal;\">\n"
"<p align=\"center\" style=\" margin-top:0px; margin-bottom:0px; margin-left:0px; margin-right:0px; -qt-block-indent:0; text-indent:0px;\">Cliquez sur une journée pour avoir les scores de celle-ci !</p></body></html>"))
        self.titre1.setHtml(_translate("La_Ligue_1_Uber_Eats",
                                             "<!DOCTYPE HTML PUBLIC \"-//W3C//DTD HTML 4.0//EN\" \"http://www.w3.org/TR/REC-html40/strict.dtd\">\n"
                                             "<html><head><meta name=\"qrichtext\" content=\"1\" /><style type=\"text/css\">\n"
                                             "p, li { white-space: pre-wrap; }\n"
                                             "</style></head><body style=\" font-family:\'Times New Roman\'; font-size:26pt; font-weight:72; font-style:normal;\">\n"
                                             "<p align=\"center\" style=\" margin-top:0px; margin-bottom:0px; margin-left:0px; margin-right:0px; -qt-block-indent:0; text-indent:0px;\">La Ligue 1 UberEats</p></body></html>"))
        self.quitter.setText(_translate("La_Ligue_1_Uber_Eats", "Quitter"))
        self.jour1.setText(_translate("La_Ligue_1_Uber_Eats", "1"))
        self.jour2.setText(_translate("La_Ligue_1_Uber_Eats", "2"))
        self.jour3.setText(_translate("La_Ligue_1_Uber_Eats", "3"))
        self.jour4.setText(_translate("La_Ligue_1_Uber_Eats", "4"))
        self.jour5.setText(_translate("La_Ligue_1_Uber_Eats", "5"))
        self.jour6.setText(_translate("La_Ligue_1_Uber_Eats", "6"))
        self.jour7.setText(_translate("La_Ligue_1_Uber_Eats", "7"))
        self.jour8.setText(_translate("La_Ligue_1_Uber_Eats", "8"))
        self.jour9.setText(_translate("La_Ligue_1_Uber_Eats", "9"))
        self.jour10.setText(_translate("La_Ligue_1_Uber_Eats", "10"))
        self.jour11.setText(_translate("La_Ligue_1_Uber_Eats", "11"))
        self.jour12.setText(_translate("La_Ligue_1_Uber_Eats", "12"))
        self.jour13.setText(_translate("La_Ligue_1_Uber_Eats", "13"))
        self.jour14.setText(_translate("La_Ligue_1_Uber_Eats", "14"))
        self.jour15.setText(_translate("La_Ligue_1_Uber_Eats", "15"))
        self.jour16.setText(_translate("La_Ligue_1_Uber_Eats", "16"))
        self.jour17.setText(_translate("La_Ligue_1_Uber_Eats", "17"))
        self.jour18.setText(_translate("La_Ligue_1_Uber_Eats", "18"))
        self.jour19.setText(_translate("La_Ligue_1_Uber_Eats", "19"))
        self.jour20.setText(_translate("La_Ligue_1_Uber_Eats", "20"))
        self.jour21.setText(_translate("La_Ligue_1_Uber_Eats", "21"))
        self.jour22.setText(_translate("La_Ligue_1_Uber_Eats", "22"))
        self.jour23.setText(_translate("La_Ligue_1_Uber_Eats", "23"))
        self.jour24.setText(_translate("La_Ligue_1_Uber_Eats", "24"))
        self.jour25.setText(_translate("La_Ligue_1_Uber_Eats", "25"))
        self.jour26.setText(_translate("La_Ligue_1_Uber_Eats", "26"))
        self.jour27.setText(_translate("La_Ligue_1_Uber_Eats", "27"))
        self.jour28.setText(_translate("La_Ligue_1_Uber_Eats", "28"))
        self.jour29.setText(_translate("La_Ligue_1_Uber_Eats", "29"))
        self.jour30.setText(_translate("La_Ligue_1_Uber_Eats", "30"))
        self.jour31.setText(_translate("La_Ligue_1_Uber_Eats", "31"))
        self.jour32.setText(_translate("La_Ligue_1_Uber_Eats", "32"))
        self.jour33.setText(_translate("La_Ligue_1_Uber_Eats", "33"))
        self.jour34.setText(_translate("La_Ligue_1_Uber_Eats", "34"))
        self.jour35.setText(_translate("La_Ligue_1_Uber_Eats", "35"))
        self.jour36.setText(_translate("La_Ligue_1_Uber_Eats", "36"))
        self.jour37.setText(_translate("La_Ligue_1_Uber_Eats", "37"))
        self.jour38.setText(_translate("La_Ligue_1_Uber_Eats", "38"))


if __name__ == "__main__":
    app = QtWidgets.QApplication(sys.argv)
    principale_ihm = QtWidgets.QMainWindow()
    ui = Ui_La_Ligue_1_Uber_Eats()
    ui.setupUi(principale_ihm)
    principale_ihm.show()
    sys.exit(app.exec_())

## Nombre de points = nombre de buts (on ne veut pas)
## Le nombre de points final de chaque club ne s'actualise pas correctement
## Afficher photo Ligue 1 plus bas ?